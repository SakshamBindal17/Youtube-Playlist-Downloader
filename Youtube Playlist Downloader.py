import csv
import yt_dlp
import time
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os

def download_video(url, video_number, title, download_folder):
    # Retry logic
    retries = 3
    for attempt in range(retries):
        try:
            # Define the file path based on video title and number
            file_path = os.path.join(download_folder, f'{video_number}. {title}.%(ext)s')

            # Check if the file already exists to skip downloading
            if os.path.exists(file_path.replace('%(ext)s', 'mp4')) or os.path.exists(file_path.replace('%(ext)s', 'webm')):
                print(f"Skipping download for {title}, already exists.")
                return True  # Skip downloading if the file exists

            ydl_opts = {
                'outtmpl': file_path,  # Use the path where the video should be downloaded
                'format': 'bestvideo[height<=1080]+bestaudio/best[height<=1080]',  # Prefer 1080p, fallback to best lower quality
                'noplaylist': True,  # To download only one video, not the entire playlist
                'quiet': True,  # Suppress unnecessary output
                'progress_hooks': [my_hook],
                'user_agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36',
            }

            with yt_dlp.YoutubeDL(ydl_opts) as ydl:
                result = ydl.extract_info(url, download=True)
                print(f"Downloaded {video_number}. {result['title']} in {result['format_id']} quality.")
            return True  # Return True if download is successful
        except yt_dlp.utils.DownloadError as e:
            print(f"Error downloading {url}: {e}. Retrying ({attempt+1}/{retries})...")
            time.sleep(3)  # Wait before retrying
        except Exception as e:
            print(f"Unexpected error: {e}")
            break  # Exit loop if the error is not related to the download
    return False  # Return False if download failed after retries

def my_hook(d):
    if d['status'] == 'finished':
        print(f"\nDownload complete: {d['filename']}")

# Prompt user for the file path of the CSV file
csv_file = input("Please enter the full path of the CSV file: ").strip()

# Remove surrounding double quotes if present
csv_file = csv_file.strip('"')

# Check if the file exists
if not os.path.exists(csv_file):
    print(f"Error: The file '{csv_file}' does not exist.")
    exit()

# Generate the Excel file name based on the CSV file name
excel_file = os.path.splitext(csv_file)[0] + '.xlsx'

# Create a folder in the CSV file's directory with the same name as the CSV file (excluding extension)
csv_dir = os.path.dirname(csv_file)
folder_name = os.path.splitext(os.path.basename(csv_file))[0]
download_folder = os.path.join(csv_dir, folder_name)

# Check if the folder exists, create it if not
if not os.path.exists(download_folder):
    os.makedirs(download_folder)

failed_videos = []  # List to store videos that failed to download

# Create a list to store rows for the new CSV file
rows_to_save = []

with open(csv_file, 'r') as file:
    reader = csv.DictReader(file)
    headers = reader.fieldnames
    for idx, row in enumerate(reader, start=1):
        title = row['Title']
        url = row['URL']
        upload_date = row['Upload Date']
        views = row['Views']

        print(f"Downloading video {idx}: {title} (URL: {url})")
        if not download_video(url, idx, title, download_folder):
            failed_videos.append((idx, title))  # Add to failed list if download failed
            row['Failed'] = 'Yes'  # Mark the row as failed in the data
        else:
            row['Failed'] = 'No'

        rows_to_save.append(row)  # Add row data to list for saving

# Now create a new Excel workbook
wb = Workbook()
ws = wb.active
ws.append(headers + ['Failed'])  # Add headers including the 'Failed' column

# Add the data rows and apply red color formatting to failed downloads
for row in rows_to_save:
    if row['Failed'] == 'Yes':
        # Apply red fill to the failed rows
        ws.append([row['Title'], row['URL'], row['Upload Date'], row['Views'], row['Failed']])
        row_idx = ws.max_row
        for cell in ws[row_idx]:
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    else:
        ws.append([row['Title'], row['URL'], row['Upload Date'], row['Views'], row['Failed']])

# Save the workbook as the Excel file
wb.save(excel_file)

# Delete the original CSV file after creating the Excel file
if os.path.exists(csv_file):
    os.remove(csv_file)
    print(f"Original CSV file {csv_file} deleted.")

# Print failed videos
if failed_videos:
    print("\nThe following videos failed to download:")
    for video_number, title in failed_videos:
        print(f"Video {video_number}: {title}")
else:
    print("\nAll videos downloaded successfully.")
